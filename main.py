""" Shopee Tool """
import sys
import os
from datetime import date
from time import sleep

import tkinter as tk
from tkinter import filedialog

from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl import load_workbook

root = tk.Tk()
root.withdraw()

DATE_TODAY = date.today()
sys.stdout.write(f"今天是{DATE_TODAY}\n")

# 路徑懶得處理 不要有中文
# 路徑懶得處理 不要有中文
# 路徑懶得處理 不要有中文

sys.stdout.write("請選擇模板\n")
TEMPLATE_FILE_NAME = filedialog.askopenfilename()
if TEMPLATE_FILE_NAME == '':
    sys.stderr.write("你沒有選擇「模板」，工具將在5秒後自動退出")
    sys.stderr.flush()
    sleep(5)
    sys.exit()
sys.stdout.write(f"模板在：{TEMPLATE_FILE_NAME}\n")

sys.stdout.write("請選擇報表存放的資料夾\n")
ORDERS_FOLDER = filedialog.askdirectory()
if ORDERS_FOLDER == '':
    sys.stderr.write("你沒有選擇「報表存放的資料夾」，工具將在5秒後自動退出")
    sys.stderr.flush()
    sleep(5)
    sys.exit()
sys.stdout.write(f"報表存放在：{ORDERS_FOLDER}\n")

sys.stdout.write("請選擇報表\n")
TODAY_ORDERS_FILE_PATH = filedialog.askopenfilename()
if TODAY_ORDERS_FILE_PATH == '':
    sys.stderr.write("你沒有選擇「報表」，工具將在5秒後自動退出")
    sys.stderr.flush()
    sleep(5)
    sys.exit()
sys.stdout.write(f"報表在：{TODAY_ORDERS_FILE_PATH}\n")

# 今日訂單存放位置
SAVE_FOLDER_PATH = os.path.join(ORDERS_FOLDER, str(DATE_TODAY))
if not os.path.exists(SAVE_FOLDER_PATH):
    os.mkdir(SAVE_FOLDER_PATH)

os.chdir(ORDERS_FOLDER)

# 讀取報表
wb = load_workbook(TODAY_ORDERS_FILE_PATH)

# 指定第一個工作頁
orders_sheet = wb[wb.sheetnames[0]]

# 自訂欄位順序
sort_target_cols_lst = [
    1, 39, 34, 33, 6, 12, 13, 14, 16, 18, 8, 24, 25, 29, 30, 27, 32, 11, 49, 50, 2, 40
]

consumer_lst = [[0 for _ in range(len(sort_target_cols_lst))] for _ in range(orders_sheet.max_row)]

# 欄位標題 放進去
for col, _ in enumerate(sort_target_cols_lst):
    consumer_lst[0][col] = orders_sheet.cell(1, sort_target_cols_lst[col]).value

# 訂單資料
for row in range(orders_sheet.max_row):

    # 只要 "待出貨" 訂單
    # 直接 continue 所以這行都是空的
    if orders_sheet.cell(row + 1, 2).value != "待出貨":
        continue

    # 按照我要的順序塞進 list (順序 與 template 的欄位順序相關)
    for col, _ in enumerate(sort_target_cols_lst):
        consumer_lst[row][col] = orders_sheet.cell(row + 1, sort_target_cols_lst[col]).value

wb.close()

# 用到 6 欄 A ~ F
title_lst = ['A', 'B', 'C', 'D', 'E', 'F']

# 中等粗細的框線 黑色
medium = Side(border_style="medium", color="000000")

# 第 0 列 是欄位標題
# 從第 1 列開始
ROW = 1
while ROW < len(consumer_lst):
    # 【待出貨】以外的訂單
    # 前面塞資料時，直接略過，所以都沒有資料
    # 這邊也略過不處理，直接進到下一筆訂單
    if not consumer_lst[ROW][0]:
        ROW += 1
        continue

    # 讀取模板
    # sole_order = load_workbook(os.path.join(TEMPLATE_PATH, TEMPLATE_FILE_NAME))
    sole_order = load_workbook(TEMPLATE_FILE_NAME)

    # 指定第一個工作頁
    current_order = sole_order[sole_order.sheetnames[0]]

    # 訂單編號 收件者姓名 收件者電話 收件者地址 訂單日期 出貨分店(這欄手動添加)
    for i in range(5):
        current_order[f'{title_lst[i]}2'] = consumer_lst[ROW][i]
        current_order[f'{title_lst[i]}2'].alignment = Alignment(vertical="center")
        current_order[f'{title_lst[i]}2'].font = Font(name='微軟正黑體')

    current_order['D2'].alignment = Alignment(wrap_text=True,
                                              horizontal="center",
                                              vertical="center")

    current_order['F2'] = "44"
    current_order['F2'].font = Font(name='Calibri', color="FF0000", bold=True)
    current_order['F2'].alignment = Alignment(horizontal="center", vertical="center")

    # 蝦皮補貼金額 蝦幣折抵 銀行信用卡活動折抵 賣場優惠券 優惠券 買家支付運費
    for i in range(5, 11):
        current_order[f'{title_lst[i-5]}4'] = consumer_lst[ROW][i]
        current_order[f'{title_lst[i-5]}4'].font = Font(name='Calibri')
        current_order[f'{title_lst[i-5]}4'].alignment = Alignment(horizontal="right",
                                                                  vertical="center")

    # 商品名稱 商品選項名稱 商品選項貨號 數量 商品活動價格 蝦皮促銷組合折扣:促銷組合標籤
    SAME_ORDER_ROW = ROW
    SAME_ORDER_COUNT = 0
    cur_order_num = consumer_lst[ROW][0]

    # 統計一次買多少件商品
    if SAME_ORDER_ROW + 1 <= len(consumer_lst):
        while consumer_lst[SAME_ORDER_ROW + 1][0] == cur_order_num:
            SAME_ORDER_COUNT += 1
            SAME_ORDER_ROW += 1
            if SAME_ORDER_ROW + 1 >= len(consumer_lst):
                break

    GOODS_INDEX_BEGIN = 6
    GOODS_INDEX_END = 6 + SAME_ORDER_COUNT

    for g in range(GOODS_INDEX_BEGIN, GOODS_INDEX_END + 1):
        for i in range(11, 17):
            current_order[f'{title_lst[i-11]}{g}'] = consumer_lst[ROW + g - 6][i]
            current_order[f'{title_lst[i-11]}{g}'].alignment = Alignment(vertical="center")
            current_order[f'{title_lst[i-11]}{g}'].border = Border(top=medium,
                                                                   left=medium,
                                                                   right=medium,
                                                                   bottom=medium)

    for g in range(GOODS_INDEX_BEGIN, GOODS_INDEX_END + 1):
        current_order[f'A{g}'] = current_order[f'A{g}'].value.replace(" ToysRUs玩具反斗城", "")
        current_order[f'A{g}'].alignment = Alignment(wrap_text=True)

        current_order[f'C{g}'].alignment = Alignment(horizontal="center", vertical="center")

        current_order[f'D{g}'].font = Font(color="FF0000", bold=True)
        current_order[f'D{g}'].alignment = Alignment(horizontal="center", vertical="center")

        current_order[f'E{g}'].alignment = Alignment(horizontal="right", vertical="center")

        current_order[f'F{g}'].alignment = Alignment(horizontal="center", vertical="center")

    # 買家總支付金額
    current_order[f'A{GOODS_INDEX_END + 1}'] = consumer_lst[0][17]
    current_order[f'A{GOODS_INDEX_END + 1}'].alignment = Alignment(vertical="center")
    current_order[f'A{GOODS_INDEX_END + 1}'].fill = PatternFill(start_color="FFFF00",
                                                                fill_type="solid")

    current_order[f'E{GOODS_INDEX_END + 1}'] = consumer_lst[ROW][17]
    current_order[f'E{GOODS_INDEX_END + 1}'].alignment = Alignment(horizontal="right",
                                                                   vertical="center")

    for i in range(6):
        current_order[f'{title_lst[i]}{GOODS_INDEX_END + 1}'].border = Border(top=medium,
                                                                              left=medium,
                                                                              right=medium,
                                                                              bottom=medium)

    # 買家備註
    if consumer_lst[ROW][18]:
        current_order[f'A{GOODS_INDEX_END + 3}'] = consumer_lst[0][18]
        current_order[f'A{GOODS_INDEX_END + 3}'].font = Font(color="FFFFFF", bold=True)
        current_order[f'A{GOODS_INDEX_END + 3}'].fill = PatternFill(start_color="C00000",
                                                                    fill_type="solid")
        current_order[f'A{GOODS_INDEX_END + 3}'].border = Border(top=medium,
                                                                 left=medium,
                                                                 right=medium,
                                                                 bottom=medium)

        current_order[f'A{GOODS_INDEX_END + 4}'] = consumer_lst[ROW][18]
        current_order[f'A{GOODS_INDEX_END + 4}'].font = Font(name='微軟正黑體')
        current_order[f'A{GOODS_INDEX_END + 4}'].alignment = Alignment(vertical="center")
        current_order[f'A{GOODS_INDEX_END + 4}'].border = Border(top=medium,
                                                                 left=medium,
                                                                 right=medium,
                                                                 bottom=medium)

    # 備註
    if consumer_lst[ROW][19]:
        current_order[f'C{GOODS_INDEX_END + 3}'] = consumer_lst[0][19]
        current_order[f'C{GOODS_INDEX_END + 3}'].font = Font(color="FFFFFF")
        current_order[f'C{GOODS_INDEX_END + 3}'].fill = PatternFill(start_color="C00000",
                                                                    fill_type="solid")
        current_order[f'C{GOODS_INDEX_END + 3}'].border = Border(top=medium,
                                                                 left=medium,
                                                                 right=medium,
                                                                 bottom=medium)

        current_order[f'C{GOODS_INDEX_END + 4}'] = consumer_lst[ROW][19]
        current_order[f'C{GOODS_INDEX_END + 4}'].font = Font(name='微軟正黑體')
        current_order[f'C{GOODS_INDEX_END + 4}'].alignment = Alignment(vertical="center")
        current_order[f'C{GOODS_INDEX_END + 4}'].border = Border(top=medium,
                                                                 left=medium,
                                                                 right=medium,
                                                                 bottom=medium)

    # 寄出方式
    SHIPPING_METHOD = '全家' if consumer_lst[ROW][-1] == '全家' else '宅配'

    # 訂單成立時間
    CURRENT_ORDER_DATE = str(consumer_lst[ROW][4]).split(" ", maxsplit=1)[0]

    # 買家姓名
    CONSUMER_NAME = str(consumer_lst[ROW][1]).replace("*", "x")

    CURRENT_ORDER_FILE_NAME = f'Shopee_{CURRENT_ORDER_DATE}_{SHIPPING_METHOD}_\
{str(consumer_lst[ROW][0])}_{CONSUMER_NAME}.xlsx'

    SAVE_FILE_PATH = os.path.join(SAVE_FOLDER_PATH, CURRENT_ORDER_FILE_NAME)
    sole_order.save(SAVE_FILE_PATH)
    sole_order.close()

    ROW += SAME_ORDER_COUNT + 1
